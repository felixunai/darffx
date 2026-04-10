"""
Router: /apuracao

Lei 14.754/2023: apuração ANUAL, alíquota fixa 15%, sem DARF mensal.

Endpoints:
  POST /apuracao/upload              — processa PDF e atualiza apuração anual
  GET  /apuracao/anual/              — lista apurações anuais do usuário
  GET  /apuracao/anual/{ano}         — detalhe anual com breakdown mensal
  GET  /apuracao/{id}                — detalhe de um mês (para drill-down)
  GET  /apuracao/anual/{ano}/pdf     — relatório PDF anual
  PATCH /apuracao/anual/{ano}/pago   — marca DARF anual como pago
  DELETE /apuracao/{id}              — remove mês (permite reprocessar)
  DELETE /apuracao/anual/{ano}       — remove todo o ano
"""
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Request
from fastapi.responses import StreamingResponse
from slowapi import Limiter
from slowapi.util import get_remote_address

limiter = Limiter(key_func=get_remote_address)

MAX_CSV_BYTES = 10 * 1024 * 1024  # 10 MB
from sqlalchemy.orm import Session
from datetime import datetime
from collections import defaultdict
import uuid
from io import BytesIO

from ..services.parser_csv_avatrade import parse_csv_avatrade
from ..services.calculo_ir import buscar_ptax_paralelo, calcular_ir_mensal, calcular_ir_anual
from ..services.gerador_pdf import gerar_relatorio_pdf
from ..models.database import Apuracao, ApuracaoAnual, Operacao, User, PtaxCache
from ..deps import get_db, get_current_user, ADMIN_EMAIL

router = APIRouter(prefix="/apuracao", tags=["apuracao"])


# ── UPLOAD ────────────────────────────────────────────────────────────────────

@router.post("/upload")
@limiter.limit("10/minute")
async def upload_extrato(
    request: Request,
    arquivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    if not arquivo.filename.lower().endswith(".csv"):
        raise HTTPException(400, "Envie um arquivo CSV.")

    _verificar_plano(usuario, db)

    conteudo = await arquivo.read()

    if len(conteudo) > MAX_CSV_BYTES:
        raise HTTPException(413, "Arquivo muito grande. Tamanho máximo: 10 MB.")

    if len(conteudo) == 0:
        raise HTTPException(400, "Arquivo vazio.")

    try:
        operacoes = parse_csv_avatrade(BytesIO(conteudo))
    except Exception as e:
        raise HTTPException(422, f"Erro ao ler o CSV: {str(e)}")

    if not operacoes:
        raise HTTPException(422, "Nenhuma operação encontrada no arquivo.")

    # Agrupa por mês/ano
    por_mes: dict = defaultdict(list)
    for op in operacoes:
        chave = (op.data.year, op.data.month)
        por_mes[chave].append(op)

    tem_closed = any(op.tipo == "CLOSED" for ops in por_mes.values() for op in ops)
    if not tem_closed:
        raise HTTPException(422, "Nenhuma operação CLOSED encontrada.")

    # Lê carry forward de anos já existentes
    apuracoes_existentes = (
        db.query(Apuracao)
        .filter_by(user_id=usuario.id)
        .order_by(Apuracao.ano, Apuracao.mes)
        .all()
    )
    acumulado_brl = _calcular_carry_forward(apuracoes_existentes)

    # Para plano free, limita a 2 meses totais
    eh_free = (usuario.plano not in ("pago", "admin") and usuario.email != ADMIN_EMAIL)
    count_existente = db.query(Apuracao).filter_by(user_id=usuario.id).count()
    slots_free = max(0, 2 - count_existente) if eh_free else None

    todos_novos = [
        (ano, mes) for (ano, mes) in sorted(por_mes.keys())
        if not db.query(Apuracao).filter_by(user_id=usuario.id, mes=mes, ano=ano).first()
    ]
    meses_permitidos = set(todos_novos[:slots_free] if slots_free is not None else todos_novos)
    meses_limitados  = len(todos_novos) > len(meses_permitidos)

    meses_novos_keys = list(meses_permitidos)
    ptax_cache = await _buscar_ptax_com_cache(db, [(mes, ano) for ano, mes in meses_novos_keys])

    # Processa meses em ordem cronológica (sequencial para carry forward)
    meses_novos: dict = defaultdict(list)   # ano → [ResultadoMensal]
    for (ano, mes), ops in sorted(por_mes.items()):
        if (ano, mes) not in meses_permitidos and (ano, mes) in set(todos_novos):
            continue  # mês novo mas fora do limite free — pula

        existente = db.query(Apuracao).filter_by(
            user_id=usuario.id, mes=mes, ano=ano
        ).first()
        if existente:
            # Já processado — só acumula carry
            if existente.ganho_brl < 0:
                acumulado_brl += abs(existente.ganho_brl)
            elif existente.ganho_brl > 0:
                acumulado_brl = max(0.0, acumulado_brl - existente.ganho_brl)
            continue

        ptax = ptax_cache.get((mes, ano), 0.0)
        resultado = calcular_ir_mensal(ops, ptax, mes, ano, carry_fwd_brl=acumulado_brl)

        apuracao = Apuracao(
            id=str(uuid.uuid4()),
            user_id=usuario.id,
            mes=mes, ano=ano,
            ganhos_usd=resultado.ganhos_usd,
            perdas_usd=resultado.perdas_usd,
            custos_usd=resultado.custos_usd,
            ganho_usd=resultado.ganho_usd,
            ptax=ptax,
            ganho_brl=resultado.ganho_brl,
            carry_fwd_brl=resultado.carry_fwd_brl,
            base_ir_brl=resultado.base_tributavel_brl,
            aliquota=resultado.aliquota,
            imposto_brl=resultado.imposto_brl,
            tem_day_trade=resultado.tem_day_trade,
            depositos_usd=resultado.depositos_usd,
            saques_usd=resultado.saques_usd,
            vencimento_darf=resultado.vencimento_darf,
        )
        db.add(apuracao)

        for op in ops:
            if op.tipo not in ("CLOSED", "OPENED", "DEPOSIT", "WITHDRAWAL"):
                continue
            db.add(Operacao(
                id=str(uuid.uuid4()),
                apuracao_id=apuracao.id,
                adj_no=op.adj_no,
                data=op.data,
                tipo=op.tipo,
                descricao=op.descricao,
                valor_usd=op.valor_usd,
            ))

        meses_novos[ano].append(resultado)

        # Atualiza carry forward
        if resultado.ganho_brl < 0:
            acumulado_brl += abs(resultado.ganho_brl)
        elif resultado.ganho_brl > 0:
            acumulado_brl = max(0.0, acumulado_brl - resultado.ganho_brl)

    db.commit()

    # Cria/atualiza ApuracaoAnual para cada ano afetado
    anos_afetados = set(meses_novos.keys())
    for ano in anos_afetados:
        _recalcular_anual(db, usuario.id, ano)

    db.commit()

    # Retorna as apurações anuais atualizadas
    anuais = (
        db.query(ApuracaoAnual)
        .filter_by(user_id=usuario.id)
        .order_by(ApuracaoAnual.ano.desc())
        .all()
    )
    total_meses = sum(len(v) for v in meses_novos.values())
    return {
        "total": total_meses,
        "meses_limitados": meses_limitados,
        "apuracoes_anuais": [_anual_to_dict(a, usuario) for a in anuais],
    }


# ── LISTAGEM ANUAL ────────────────────────────────────────────────────────────

@router.get("/anual/")
def listar_anuais(
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    anuais = (
        db.query(ApuracaoAnual)
        .filter_by(user_id=usuario.id)
        .order_by(ApuracaoAnual.ano.desc())
        .all()
    )
    # Se não existem registros anuais mas existem mensais, recalcula
    if not anuais:
        anos = {a.ano for a in db.query(Apuracao).filter_by(user_id=usuario.id).all()}
        for ano in sorted(anos):
            _recalcular_anual(db, usuario.id, ano)
        db.commit()
        anuais = (
            db.query(ApuracaoAnual)
            .filter_by(user_id=usuario.id)
            .order_by(ApuracaoAnual.ano.desc())
            .all()
        )
    return [_anual_to_dict(a, usuario) for a in anuais]


@router.get("/anual/{ano}")
def detalhe_anual(
    ano: int,
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    anual = db.query(ApuracaoAnual).filter_by(user_id=usuario.id, ano=ano).first()
    if not anual:
        raise HTTPException(404, "Apuração anual não encontrada.")

    meses = (
        db.query(Apuracao)
        .filter_by(user_id=usuario.id, ano=ano)
        .order_by(Apuracao.mes)
        .all()
    )
    return {**_anual_to_dict(anual, usuario), "meses": [_mensal_to_dict(m) for m in meses]}


@router.patch("/anual/{ano}/pago")
def marcar_anual_pago(
    ano: int,
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    anual = db.query(ApuracaoAnual).filter_by(user_id=usuario.id, ano=ano).first()
    if not anual:
        raise HTTPException(404, "Apuração anual não encontrada.")
    anual.darf_pago = True
    db.commit()
    return _anual_to_dict(anual)


@router.patch("/anual/{ano}/pendente")
def marcar_anual_pendente(
    ano: int,
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    anual = db.query(ApuracaoAnual).filter_by(user_id=usuario.id, ano=ano).first()
    if not anual:
        raise HTTPException(404, "Apuração anual não encontrada.")
    anual.darf_pago = False
    db.commit()
    return _anual_to_dict(anual)


@router.delete("/anual/{ano}")
def deletar_anual(
    ano: int,
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    anual = db.query(ApuracaoAnual).filter_by(user_id=usuario.id, ano=ano).first()
    if anual:
        db.delete(anual)
    meses = db.query(Apuracao).filter_by(user_id=usuario.id, ano=ano).all()
    for m in meses:
        db.query(Operacao).filter_by(apuracao_id=m.id).delete()
        db.delete(m)
    db.commit()
    return {"ok": True}


# ── DETALHE MENSAL (drill-down) ───────────────────────────────────────────────

@router.get("/")
def listar_mensais(
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    """Mantido para compatibilidade — retorna meses ordenados."""
    apuracoes = (
        db.query(Apuracao)
        .filter_by(user_id=usuario.id)
        .order_by(Apuracao.ano.desc(), Apuracao.mes.desc())
        .all()
    )
    return [_mensal_to_dict(a) for a in apuracoes]


@router.get("/{apuracao_id}")
def detalhe_mensal(
    apuracao_id: str,
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    apuracao = db.query(Apuracao).filter_by(id=apuracao_id, user_id=usuario.id).first()
    if not apuracao:
        raise HTTPException(404, "Apuração não encontrada.")
    return _mensal_to_dict(apuracao, incluir_operacoes=True)


@router.delete("/{apuracao_id}")
def deletar_mensal(
    apuracao_id: str,
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    apuracao = db.query(Apuracao).filter_by(id=apuracao_id, user_id=usuario.id).first()
    if not apuracao:
        raise HTTPException(404, "Apuração não encontrada.")
    ano = apuracao.ano
    db.query(Operacao).filter_by(apuracao_id=apuracao_id).delete()
    db.delete(apuracao)
    db.commit()
    _recalcular_anual(db, usuario.id, ano)
    db.commit()
    return {"ok": True}


@router.patch("/{apuracao_id}/ptax")
def atualizar_ptax(
    apuracao_id: str,
    ptax: float,
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    apuracao = db.query(Apuracao).filter_by(id=apuracao_id, user_id=usuario.id).first()
    if not apuracao:
        raise HTTPException(404, "Apuração não encontrada.")
    apuracao.ptax = ptax
    apuracao.ganho_brl = round(apuracao.ganho_usd * ptax, 2)
    carry = apuracao.carry_fwd_brl or 0.0
    base  = round(max(0.0, apuracao.ganho_brl - carry), 2) if apuracao.ganho_brl > 0 else 0.0
    apuracao.base_ir_brl = base
    apuracao.imposto_brl = round(base * apuracao.aliquota, 2)
    db.commit()
    _recalcular_anual(db, usuario.id, apuracao.ano)
    db.commit()
    return _mensal_to_dict(apuracao)


@router.get("/anual/{ano}/xlsx")
def download_xlsx(
    ano: int,
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    """Exporta apuração anual como planilha Excel (.xlsx)."""
    anual = db.query(ApuracaoAnual).filter_by(user_id=usuario.id, ano=ano).first()
    if not anual:
        raise HTTPException(404, "Apuração anual não encontrada.")
    if not _is_desbloqueado(anual, usuario):
        raise HTTPException(403, "Relatório bloqueado. Faça o upgrade para exportar.")

    meses = (
        db.query(Apuracao)
        .filter_by(user_id=usuario.id, ano=ano)
        .order_by(Apuracao.mes)
        .all()
    )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(503, "Dependência de exportação não instalada.")

    MESES_NOME = [
        "Janeiro","Fevereiro","Março","Abril","Maio","Junho",
        "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro",
    ]
    BG_DARK = "0A0E17"; BG_CARD = "131929"; BG_ALT = "1A2235"
    GREEN   = "00E5A0"; RED = "FF4D6D";     WARN  = "FFB347"
    TEXT    = "E8EDF5"; MUTED = "8899AA";   BLACK = "000000"

    wb = Workbook()
    ws = wb.active
    ws.title = f"DarfFX {ano}"

    # Título
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = f"DarfFX — Apuração IR Forex {ano}  ·  Lei 14.754/2023  ·  Alíquota 15%"
    c.font = Font(bold=True, color=GREEN, size=13, name="Calibri")
    c.fill = PatternFill(start_color=BG_DARK, end_color=BG_DARK, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Cabeçalho
    headers = [
        "Mês", "Ganhos (USD)", "Perdas (USD)", "Líquido (USD)",
        "PTAX (R$)", "Líquido (BRL)", "Prej. Comp. (BRL)",
        "Base IR (BRL)", "Alíquota", "Imposto (BRL)",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, color=BLACK, name="Calibri", size=10)
        c.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Dados mensais
    for i, m in enumerate(meses, 3):
        bg = BG_CARD if i % 2 == 1 else BG_ALT
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        ganho_usd = round(m.ganho_usd or 0, 2)
        ganho_brl = round(m.ganho_brl or 0, 2)
        imposto   = round(m.imposto_brl or 0, 2)

        values = [
            MESES_NOME[m.mes - 1],
            round(m.ganhos_usd or 0, 2),
            round(m.perdas_usd or 0, 2),
            ganho_usd,
            round(m.ptax or 0, 4),
            ganho_brl,
            round(m.carry_fwd_brl or 0, 2),
            round(m.base_ir_brl or 0, 2),
            f"{int((m.aliquota or 0.15) * 100)}%",
            imposto,
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.fill = fill
            if col == 4:
                c.font = Font(color=(GREEN if ganho_usd >= 0 else RED), bold=True, name="Calibri", size=10)
            elif col == 6:
                c.font = Font(color=(GREEN if ganho_brl >= 0 else RED), bold=True, name="Calibri", size=10)
            elif col == 10:
                c.font = Font(color=(WARN if imposto > 0 else GREEN), bold=True, name="Calibri", size=10)
            else:
                c.font = Font(color=(TEXT if col == 1 else MUTED), name="Calibri", size=10)
            c.alignment = Alignment(horizontal=("left" if col == 1 else "center"), vertical="center")

    # Linha de totais
    total_row = len(meses) + 3
    ws.row_dimensions[total_row].height = 22
    totals_fill = PatternFill(start_color=BG_DARK, end_color=BG_DARK, fill_type="solid")
    totals = {
        1: "TOTAL",
        2: round(sum(m.ganhos_usd or 0 for m in meses), 2),
        3: round(sum(m.perdas_usd or 0 for m in meses), 2),
        4: round(sum(m.ganho_usd  or 0 for m in meses), 2),
        6: round(sum(m.ganho_brl  or 0 for m in meses), 2),
        10: round(anual.imposto_brl or 0, 2),
    }
    for col in range(1, 11):
        c = ws.cell(row=total_row, column=col, value=totals.get(col, ""))
        c.fill = totals_fill
        c.font = Font(bold=True, color=GREEN, name="Calibri", size=10)
        c.alignment = Alignment(horizontal=("left" if col == 1 else "center"), vertical="center")

    # Largura das colunas
    for i, w in enumerate([16, 14, 14, 14, 10, 14, 16, 14, 9, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=darffx_{ano}.xlsx"},
    )


@router.get("/anual/{ano}/pares")
def pares_por_ano(
    ano: int,
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    """Retorna P&L agregado por instrumento/par para o ano selecionado."""
    import re as _re
    anual = db.query(ApuracaoAnual).filter_by(user_id=usuario.id, ano=ano).first()
    if not anual:
        raise HTTPException(404, "Apuração anual não encontrada.")
    if not _is_desbloqueado(anual, usuario):
        raise HTTPException(403, "Relatório bloqueado. Faça o upgrade para acessar.")

    apuracoes = db.query(Apuracao).filter_by(user_id=usuario.id, ano=ano).all()
    ids = [a.id for a in apuracoes]
    if not ids:
        return []

    ops = (
        db.query(Operacao)
        .filter(Operacao.apuracao_id.in_(ids), Operacao.tipo == "CLOSED")
        .all()
    )

    from collections import defaultdict
    pares: dict = defaultdict(lambda: {"trades": 0, "lucro_usd": 0.0})
    for op in ops:
        par = _normalizar_par(op.descricao)
        pares[par]["trades"] += 1
        pares[par]["lucro_usd"] = round(pares[par]["lucro_usd"] + (op.valor_usd or 0.0), 2)

    return sorted(
        [{"par": k, **v} for k, v in pares.items()],
        key=lambda x: x["lucro_usd"],
        reverse=True,
    )


def _normalizar_par(descricao: str) -> str:
    """Extrai nome do instrumento da descrição da AvaTrade."""
    import re
    if not descricao:
        return "Outros"
    d = descricao.strip()
    # Remove prefixos buy/sell que eventualmente aparecem
    d = re.sub(r'^(buy|sell|long|short)\s+', '', d, flags=re.IGNORECASE)
    # Pega tudo antes de " @", " at " ou parênteses (nomes como "EUR/USD @ 1.08")
    d = re.split(r'\s+@|\s+at\s+|\s*\(', d, flags=re.IGNORECASE)[0].strip()
    return d[:30] if d else "Outros"


@router.get("/{apuracao_id}/pdf")
def download_pdf(
    apuracao_id: str,
    db: Session = Depends(get_db),
    usuario: User = Depends(get_current_user),
):
    apuracao = db.query(Apuracao).filter_by(id=apuracao_id, user_id=usuario.id).first()
    if not apuracao:
        raise HTTPException(404, "Apuração não encontrada.")

    from ..services.calculo_ir import ResultadoMensal
    resultado = ResultadoMensal(
        mes=apuracao.mes, ano=apuracao.ano,
        ganho_usd=apuracao.ganho_usd, ptax=apuracao.ptax or 0,
        ganho_brl=apuracao.ganho_brl,
        carry_fwd_brl=apuracao.carry_fwd_brl or 0,
        base_tributavel_brl=apuracao.base_ir_brl or 0,
        aliquota=apuracao.aliquota, imposto_brl=apuracao.imposto_brl,
        tem_day_trade=apuracao.tem_day_trade,
        operacoes_count=len(apuracao.operacoes),
        depositos_usd=apuracao.depositos_usd or 0,
        saques_usd=apuracao.saques_usd or 0,
        vencimento_darf=apuracao.vencimento_darf,
    )
    pdf_bytes = gerar_relatorio_pdf(resultado, usuario.nome or usuario.email)
    nome = f"darffx_{apuracao.ano}_{apuracao.mes:02d}.pdf"
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={nome}"})


# ── HELPERS INTERNOS ──────────────────────────────────────────────────────────

async def _buscar_ptax_com_cache(
    db: Session,
    meses_anos: list[tuple[int, int]],
) -> dict[tuple[int, int], float]:
    """
    Retorna PTAX para cada (mes, ano) usando cache do banco.
    Apenas meses sem cache são consultados na API BCB.
    Se o BCB estiver indisponível, usa o valor mais recente do cache como fallback.
    """
    resultado: dict[tuple[int, int], float] = {}
    sem_cache: list[tuple[int, int]] = []

    for mes, ano in meses_anos:
        cached = db.query(PtaxCache).filter_by(mes=mes, ano=ano).first()
        if cached:
            resultado[(mes, ano)] = cached.ptax
        else:
            sem_cache.append((mes, ano))

    if sem_cache:
        novos = await buscar_ptax_paralelo(sem_cache)
        for (mes, ano), ptax in novos.items():
            if ptax:
                db.add(PtaxCache(mes=mes, ano=ano, ptax=ptax))
        db.commit()
        resultado.update(novos)

    # Fallback: se BCB retornou 0 para algum mês, usa PTAX mais recente do cache
    for mes, ano in meses_anos:
        if not resultado.get((mes, ano)):
            fallback = (
                db.query(PtaxCache)
                .order_by(PtaxCache.ano.desc(), PtaxCache.mes.desc())
                .first()
            )
            if fallback:
                print(
                    f"[PTAX-FALLBACK] ({mes}/{ano}) BCB indisponível — "
                    f"usando cache de {fallback.mes}/{fallback.ano}: {fallback.ptax}",
                    flush=True,
                )
                resultado[(mes, ano)] = fallback.ptax

    return resultado


def _recalcular_anual(db: Session, user_id: str, ano: int):
    """Recalcula (ou cria) o registro ApuracaoAnual a partir dos meses existentes."""
    from ..services.calculo_ir import calcular_ir_anual, ResultadoMensal

    meses_db = (
        db.query(Apuracao)
        .filter_by(user_id=user_id, ano=ano)
        .order_by(Apuracao.mes)
        .all()
    )
    if not meses_db:
        return

    # Carry forward de anos anteriores (anos com prejuízo)
    anos_anteriores = (
        db.query(ApuracaoAnual)
        .filter_by(user_id=user_id)
        .filter(ApuracaoAnual.ano < ano)
        .order_by(ApuracaoAnual.ano)
        .all()
    )
    prejuizo_acum = 0.0
    for ap in anos_anteriores:
        if ap.lucro_brl < 0:
            prejuizo_acum += abs(ap.lucro_brl)
        elif ap.lucro_brl > 0:
            prejuizo_acum = max(0.0, prejuizo_acum - ap.lucro_brl)

    # Converte registros DB em ResultadoMensal simples
    meses_resultado = [
        ResultadoMensal(
            mes=m.mes, ano=m.ano,
            ganhos_usd=m.ganhos_usd or 0,
            perdas_usd=m.perdas_usd or 0,
            custos_usd=m.custos_usd or 0,
            ganho_usd=m.ganho_usd or 0,
            ptax=m.ptax or 0,
            ganho_brl=m.ganho_brl or 0,
            carry_fwd_brl=m.carry_fwd_brl or 0,
            base_tributavel_brl=m.base_ir_brl or 0,
            aliquota=m.aliquota or 0.15,
            imposto_brl=m.imposto_brl or 0,
            tem_day_trade=m.tem_day_trade or False,
            operacoes_count=len(m.operacoes),
            depositos_usd=m.depositos_usd or 0,
            saques_usd=m.saques_usd or 0,
            vencimento_darf=m.vencimento_darf,
        )
        for m in meses_db
    ]

    resultado = calcular_ir_anual(meses_resultado, ano, prejuizo_acum)

    anual = db.query(ApuracaoAnual).filter_by(user_id=user_id, ano=ano).first()
    if not anual:
        anual = ApuracaoAnual(id=str(uuid.uuid4()), user_id=user_id, ano=ano)
        db.add(anual)

    anual.lucro_usd             = resultado.lucro_usd
    anual.lucro_brl             = resultado.lucro_brl
    anual.prejuizo_anterior_brl = resultado.prejuizo_anterior_brl
    anual.base_tributavel_brl   = resultado.base_tributavel_brl
    anual.aliquota              = resultado.aliquota
    anual.imposto_brl           = resultado.imposto_brl
    anual.depositos_usd         = resultado.depositos_usd
    anual.saques_usd            = resultado.saques_usd
    anual.vencimento_darf       = resultado.vencimento_darf

    # Vincula meses ao registro anual
    for m in meses_db:
        m.apuracao_anual_id = anual.id


def _calcular_carry_forward(apuracoes: list) -> float:
    acum = 0.0
    for a in apuracoes:
        if a.ganho_brl < 0:
            acum += abs(a.ganho_brl)
        elif a.ganho_brl > 0:
            acum = max(0.0, acum - a.ganho_brl)
    return acum


def _verificar_plano(usuario, db: Session):
    """
    Planos:
    - admin: sem restrição
    - pago: sem restrição até 31/12 do ano vigente
    - free: até 2 meses de análise — resultado fica bloqueado até pagar
    """
    if usuario.email == ADMIN_EMAIL or usuario.plano == "admin":
        return
    if usuario.plano == "pago":
        if usuario.plano_expiracao and usuario.plano_expiracao < datetime.utcnow():
            raise HTTPException(402, "Seu acesso expirou em 31/12. Renove para continuar.")
        return
    # free: permite até 2 meses (Apuracao records)
    count = db.query(Apuracao).filter_by(user_id=usuario.id).count()
    if count >= 2:
        raise HTTPException(
            402,
            "PLAN_LIMIT: Plano gratuito permite 2 meses de análise. "
            "Desbloqueie o Acesso Completo para processar mais meses."
        )


def _r2(v) -> float:
    return round(v or 0, 2)


def _is_desbloqueado(a: ApuracaoAnual, usuario) -> bool:
    """Admin e plano pago válido desbloqueiam tudo; senão verifica flag da apuração."""
    if usuario.email == ADMIN_EMAIL or usuario.plano == "admin":
        return True
    if usuario.plano == "pago":
        if usuario.plano_expiracao is None or usuario.plano_expiracao > datetime.utcnow():
            return True
    return bool(a.desbloqueado)


def _anual_to_dict(a: ApuracaoAnual, usuario=None) -> dict:
    desbloqueado = _is_desbloqueado(a, usuario) if usuario else bool(a.desbloqueado)

    base = {
        "id": a.id,
        "ano": a.ano,
        "desbloqueado": desbloqueado,
        "lucro_usd": _r2(a.lucro_usd),
        "lucro_brl": _r2(a.lucro_brl),
        "depositos_usd": _r2(a.depositos_usd),
        "saques_usd": _r2(a.saques_usd),
        "darf_pago": a.darf_pago,
        "created_at": a.created_at.isoformat() if a.created_at else None,
    }

    if desbloqueado:
        # Dados completos — apenas para usuários que pagaram
        base.update({
            "prejuizo_anterior_brl": _r2(a.prejuizo_anterior_brl),
            "base_tributavel_brl": _r2(a.base_tributavel_brl),
            "aliquota": _r2(a.aliquota),
            "imposto_brl": _r2(a.imposto_brl),
            "vencimento_darf": a.vencimento_darf.isoformat() if a.vencimento_darf else None,
        })
    else:
        # Teaser — sem imposto, sem base tributável
        base.update({
            "prejuizo_anterior_brl": None,
            "base_tributavel_brl": None,
            "aliquota": None,
            "imposto_brl": None,
            "vencimento_darf": None,
        })

    return base


def _mensal_to_dict(a: Apuracao, incluir_operacoes=False) -> dict:
    d = {
        "id": a.id,
        "mes": a.mes,
        "ano": a.ano,
        "ganhos_usd": _r2(a.ganhos_usd),
        "perdas_usd": _r2(a.perdas_usd),
        "custos_usd": _r2(a.custos_usd),
        "ganho_usd": _r2(a.ganho_usd),
        "ptax": round(a.ptax, 4) if a.ptax else None,
        "ganho_brl": _r2(a.ganho_brl),
        "carry_fwd_brl": _r2(a.carry_fwd_brl),
        "base_ir_brl": _r2(a.base_ir_brl),
        "aliquota": _r2(a.aliquota),
        "imposto_brl": _r2(a.imposto_brl),
        "tem_day_trade": a.tem_day_trade,
        "depositos_usd": _r2(a.depositos_usd),
        "saques_usd": _r2(a.saques_usd),
        "vencimento_darf": a.vencimento_darf.isoformat() if a.vencimento_darf else None,
        "operacoes_count": len(a.operacoes),
        "darf_pago": a.darf_pago,
        "created_at": a.created_at.isoformat() if a.created_at else None,
    }
    if incluir_operacoes:
        d["operacoes"] = sorted(
            [
                {
                    "adj_no": op.adj_no,
                    "data": op.data.isoformat() if op.data else None,
                    "tipo": op.tipo,
                    "descricao": op.descricao,
                    "valor_usd": _r2(op.valor_usd),
                }
                for op in a.operacoes
            ],
            key=lambda x: x["data"] or ""
        )
    return d
